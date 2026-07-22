from __future__ import annotations

from pathlib import Path
from typing import Iterable

from local_full_text_search.core.errors import ParserDependencyError
from local_full_text_search.core.task_manager import CancelToken
from local_full_text_search.models.content_block import ContentBlock
from local_full_text_search.parsers.base_parser import BaseParser


class XlsxParser(BaseParser):
    """Stream XLSX/XLSM rows into searchable blocks.

    openpyxl read-only mode is intentionally kept here because training and
    service spreadsheets can be very large. In that mode sparse rows may contain
    EmptyCell objects without row/coordinate attributes, so coordinate access
    must happen only after a real value has been confirmed.
    """

    name = "xlsx"

    def supports(self, file_path: Path) -> bool:
        return file_path.suffix.lower() in {".xlsx", ".xlsm"}

    def parse(self, file_path: Path, cancel_token: CancelToken) -> Iterable[ContentBlock]:
        try:
            from openpyxl import load_workbook
            from openpyxl.cell.read_only import EmptyCell
        except ImportError as exc:
            raise ParserDependencyError("未安装 openpyxl，无法解析 Excel") from exc
        workbook = load_workbook(filename=file_path, read_only=True, data_only=False)
        block_index = 0
        try:
            for sheet in workbook.worksheets:
                cancel_token.wait_if_paused()
                cancel_token.throw_if_cancelled()
                for row_number, row in enumerate(sheet.iter_rows(), start=1):
                    cancel_token.wait_if_paused()
                    cancel_token.throw_if_cancelled()
                    parts: list[str] = []
                    first_cell = None
                    last_cell = None
                    for cell in row:
                        # read_only 模式下稀疏表格会返回 EmptyCell；它没有 row、
                        # coordinate 等定位属性，所以必须先跳过，再访问坐标。
                        if isinstance(cell, EmptyCell):
                            continue
                        value = cell.value
                        if value is None:
                            continue
                        coordinate = getattr(cell, "coordinate", None)
                        if not coordinate:
                            continue
                        if first_cell is None:
                            first_cell = coordinate
                        last_cell = coordinate
                        text = str(value)
                        comment = getattr(getattr(cell, "comment", None), "text", None)
                        if comment:
                            text = f"{text}；批注：{comment}"
                        parts.append(f"{coordinate}={text}")
                    if not parts:
                        continue
                    yield self.make_block(
                        file_path,
                        block_index,
                        "xlsx_row",
                        f"Sheet：{sheet.title}；第 {row_number} 行",
                        " | ".join(parts),
                        sheet_name=sheet.title,
                        cell_start=first_cell,
                        cell_end=last_cell,
                    )
                    block_index += 1
        finally:
            workbook.close()
